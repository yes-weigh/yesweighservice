#!/usr/bin/env python3
"""Extract BdService workbook → JSONL for blueDartPincodes seed."""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "BdService (32).xlsx"
OUT = ROOT / "scripts" / "data" / "bluedart-pincodes.jsonl"


def yes_bool(value) -> bool:
    return str(value or "").strip().lower() in {"yes", "y", "true", "1"}


def main() -> int:
    if not SRC.exists():
        print(f"Missing {SRC}", file=sys.stderr)
        return 1
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip().upper() for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    def cell(row, key, default=""):
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return "" if v is None else v

    count = 0
    with OUT.open("w", encoding="utf-8") as fh:
        for row in rows:
            pin = str(cell(row, "CPINCODE")).strip()
            digits = "".join(ch for ch in pin if ch.isdigit())
            if len(digits) != 6:
                continue
            dp_zone = str(cell(row, "DP_ZONE")).strip().upper()
            if dp_zone not in {"A", "B", "C"}:
                dp_zone = ""
            doc = {
                "pincode": digits,
                "region": str(cell(row, "CREGION")).strip(),
                "state": str(cell(row, "CSTATE")).strip(),
                "area": str(cell(row, "CAREA")).strip(),
                "areaDesc": str(cell(row, "CAREADESC")).strip(),
                "hubCode": str(cell(row, "CSCRCD")).strip(),
                "dpService": str(cell(row, "DPSERVICE")).strip() or "No",
                "dpZone": dp_zone,
                "apxService": str(cell(row, "APXSERVICE")).strip() or "No",
                "sfcService": str(cell(row, "SFCSERVICE")).strip() or "No",
                "edlApx": yes_bool(cell(row, "EDL_APX")),
                "edlSfc": yes_bool(cell(row, "EDL_SFC")),
                "edlKm": None,
                "apxLocIb": str(cell(row, "APX_LOCIB")).strip(),
                "sfcLocIb": str(cell(row, "SFC_LOCIB")).strip(),
            }
            fh.write(json.dumps(doc, ensure_ascii=False) + "\n")
            count += 1
    print(f"Wrote {count} pins -> {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
